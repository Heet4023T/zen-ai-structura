import os
import json
import requests
import base64
import re
import datetime
from flask import Flask, request, render_template, jsonify, send_file, redirect, url_for, flash, session
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image

# ==============================================================================
# 1. CONFIGURATION & SETUP
# ==============================================================================
API_KEY = os.environ.get("GITHUB_TOKEN")
MODEL = "gpt-4o" 
API_URL = "https://models.inference.ai.azure.com/chat/completions"

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")
LAST_GENERATED_FILE = None
os.makedirs(UPLOAD_DIR, exist_ok=True)

app = Flask(__name__)
app.config['SECRET_KEY'] = 'FINAL_LAYOUT_SWITCHING_FIX_V21' 
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///users.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['PERMANENT_SESSION_LIFETIME'] = datetime.timedelta(days=30)

# ==============================================================================
# 2. DATABASE MODELS
# ==============================================================================
db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login' 

class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(150))
    email = db.Column(db.String(150), unique=True)
    password = db.Column(db.String(150))
    gender = db.Column(db.String(50))

class History(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    filename = db.Column(db.String(300))
    prompt = db.Column(db.Text)
    timestamp = db.Column(db.DateTime, default=datetime.datetime.utcnow)

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

with app.app_context():
    db.create_all()

# ==============================================================================
# 3. UTILITY FUNCTIONS
# ==============================================================================
def clean(v):
    if v is None: return ""
    s = str(v).strip()
    if s.lower() in ["null", "none", "n/a", "", "[]", "{}"]: return ""
    return s

def encode_image(image_path):
    with open(image_path, "rb") as image_file:
        return base64.b64encode(image_file.read()).decode('utf-8')

def extract_number(value):
    if not value: return 0.0
    matches = re.findall(r"(-?\d+(?:\.\d+)?)", str(value).replace(",", ""))
    return float(matches[0]) if matches else 0.0

def recalculate_math(data):
    items = data.get("items", [])
    footer = data.get("footer", {})
    is_personal_mode = (data.get("layout") == "personal")
    running_total = 0.0
    
    global_tax_pct = 0.0
    if not is_personal_mode:
        tax_str = str(footer.get("tax_summary") or "")
        rates = re.findall(r"(\d+(?:\.\d+)?)", tax_str)
        if rates:
            raw = [float(r) for r in rates if float(r) <= 50]
            if raw:
                s = sum(raw)
                global_tax_pct = s if any(abs(s - x) < 0.1 for x in [5,12,18,28]) else max(raw)
        if abs(global_tax_pct - 9.0) < 0.1: global_tax_pct = 18.0

    for item in items:
        qty = extract_number(item.get("quantity"))
        rate = extract_number(item.get("rate"))
        disc_pct = extract_number(item.get("discount_percent"))
        desc = str(item.get("particulars") or "").lower()

        if ("discount" in desc or "less" in desc) and rate > 0: rate *= -1
        if qty == 0 and rate != 0: qty = 1.0

        gross = qty * rate
        discount_amount = 0.0
        
        if not is_personal_mode:
            if disc_pct > 0:
                discount_amount = gross * (disc_pct / 100.0)
            
            taxable = gross - discount_amount
            tax_amt = taxable * (global_tax_pct / 100.0)
            final_amt = taxable + tax_amt
            item["tax_rate"] = f"{int(global_tax_pct)}%"
        else:
            final_amt = gross
            item["tax_rate"] = "0%"

        item.update({
            "quantity": qty, 
            "rate": rate, 
            "gross_amount": round(gross, 2),
            "discount_amount": round(discount_amount, 2),
            "amount": round(final_amt, 2)
        })
        running_total += final_amt

    footer["total_amount"] = round(running_total, 2)
    return data

# ==============================================================================
# 4. AI PARSING LOGIC (FIXED LAYOUT SWITCHING)
# ==============================================================================
def parse_invoice_vision(image_path, user_instruction=""):
    base64_img = encode_image(image_path)
    
    # UPDATED PROMPT: STRICTER RULES FOR PERSONAL VS BUSINESS
    prompt = f"""
    Analyze the input and extract data into JSON. 
    USER INSTRUCTION: "{user_instruction}"
    
    CRITICAL LAYOUT RULES (PRIORITY):
    1. **PERSONAL MODE (STRICT)**: If the USER INSTRUCTION contains phrases like "paid to", "bought from", "spent on", or is a simple list of items/expenses, YOU MUST SET "layout": "personal". 
       - In this mode, DO NOT extract any company names, addresses, GSTINs, or bank details. Leave the "header" object completely null.
    
    2. **BUSINESS MODE**: Only set "layout": "business" if the image clearly contains a formal invoice header, such as a distinct Company Title (e.g., "Sharma Enterprises"), "Tax Invoice", "GSTIN", or detailed address blocks.

    EXTRACTION IF BUSINESS MODE:
       - **COMPANY NAME**: Extract top title.
       - **SUBTEXT**: Extract address/GSTIN lines IMMEDIATELY BELOW the title.
       - **INVOICE NO**: Look for "Inv No", "Invoice #".
    
    JSON STRUCTURE:
    {{
      "layout": "business" OR "personal",
      "header": {{ "company_name": null, "company_subtext": null, "gstin": null, "buyer_name": null, "invoice_no": null, "date": null, "bank_details": {{ "bank_name": null, "acc_no": null, "ifsc": null }} }},
      "items": [ {{ "sn": 1, "particulars": null, "hsn_sac": null, "quantity": 0, "rate": 0, "discount_percent": 0, "amount": 0 }} ],
      "footer": {{ "tax_summary": null, "total_amount": 0, "amount_in_words": null }}
    }}
    """
    
    payload = {"model": MODEL, "messages": [{"role": "user", "content": [{"type": "text", "text": prompt}, {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{base64_img}"}}]}]}
    headers = {"Authorization": f"Bearer {API_KEY}", "Content-Type": "application/json"}
    
    r = requests.post(API_URL, headers=headers, json=payload, timeout=60)
    raw = r.json()["choices"][0]["message"]["content"]
    json_str = re.sub(r'[\x00-\x09\x0b-\x1f\x7f]', '', raw[raw.find("{"):raw.rfind("}")+1])
    return recalculate_math(json.loads(json_str, strict=False))

# ==============================================================================
# 5. EXCEL LAYOUTS (FIXED PERSONAL LAYOUT)
# ==============================================================================
def draw_box(ws, cell_range, value, font=None, align=None, fill=None, border=None):
    ws.merge_cells(cell_range)
    top_left = ws[cell_range.split(':')[0]]
    top_left.value = value
    if font: top_left.font = font
    if align: top_left.alignment = align
    if fill: top_left.fill = fill
    if border:
        for row in ws[cell_range]:
            for c in row: c.border = border

def set_outer_border(ws, cell_range):
    thick = Side(style='medium', color='000000')
    rows = list(ws[cell_range])
    for cell in rows[0]: cell.border = Border(top=thick, left=cell.border.left, right=cell.border.right, bottom=cell.border.bottom)
    for cell in rows[-1]: cell.border = Border(top=cell.border.top, left=cell.border.left, right=cell.border.right, bottom=thick)
    for row in rows: row[0].border = Border(top=row[0].border.top, left=thick, right=row[0].border.right, bottom=row[0].border.bottom)
    for row in rows: row[-1].border = Border(top=row[-1].border.top, left=row[-1].border.left, right=thick, bottom=row[-1].border.bottom)
    rows[0][0].border = Border(top=thick, left=thick, right=rows[0][0].border.right, bottom=rows[0][0].border.bottom)
    rows[0][-1].border = Border(top=thick, left=rows[0][-1].border.left, right=thick, bottom=rows[0][-1].border.bottom)
    rows[-1][0].border = Border(top=rows[-1][0].border.top, left=thick, right=rows[-1][0].border.right, bottom=thick)
    rows[-1][-1].border = Border(top=rows[-1][-1].border.top, left=rows[-1][-1].border.left, right=thick, bottom=thick)

def write_business_layout(ws, data):
    head, items, foot = data.get("header", {}), data.get("items", []), data.get("footer", {})
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right = Alignment(horizontal='right', vertical='center')
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    box_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    has_disc = any(item.get("discount_amount", 0) > 0 for item in items)
    headers = ["S.N.", "Particulars", "HSN/SAC", "Qty", "Rate"]
    keys = ["sn", "particulars", "hsn_sac", "quantity", "rate"]
    widths = [6, 35, 12, 8, 12]

    if has_disc:
        headers.extend(["Gross", "Disc"])
        keys.extend(["gross_amount", "discount_amount"])
        widths.extend([12, 10])

    headers.extend(["Tax %", "Total"])
    keys.extend(["tax_rate", "amount"])
    widths.extend([10, 15])
    num_cols = len(headers)
    last_col_let = get_column_letter(num_cols)

    ws.row_dimensions[1].height = 35 
    comp_name = clean(head.get("company_name")) or "SHARMA ENTERPRISES"
    draw_box(ws, f'A1:{last_col_let}1', comp_name, font=Font(size=22, bold=True), align=center, fill=PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid"), border=box_border)
    ws.row_dimensions[2].height = 45
    subtext = clean(head.get("company_subtext"))
    if clean(head.get("gstin")) and "GSTIN" not in subtext: subtext += f" | GSTIN: {clean(head.get('gstin'))}"
    draw_box(ws, f'A2:{last_col_let}2', subtext, align=center, border=box_border)

    mid = num_cols // 2; mid_let = get_column_letter(mid); mid_plus_1_let = get_column_letter(mid + 1)
    ws.row_dimensions[3].height = 25; ws.row_dimensions[4].height = 25
    draw_box(ws, f'A3:{mid_let}3', f"To: {clean(head.get('buyer_name'))}", font=Font(bold=True), align=left, border=box_border)
    draw_box(ws, f'{mid_plus_1_let}3:{last_col_let}3', f"Inv No: {clean(head.get('invoice_no'))}", font=Font(bold=True), align=center, border=box_border)
    draw_box(ws, f'A4:{mid_let}4', clean(head.get('buyer_address')) or "Address", align=left, border=box_border)
    draw_box(ws, f'{mid_plus_1_let}4:{last_col_let}4', f"Date: {clean(head.get('date'))}", align=center, border=box_border)

    ws.row_dimensions[5].height = 25
    bank = head.get("bank_details", {})
    draw_box(ws, 'A5:B5', "Bank Details:", font=Font(bold=True), align=left, border=box_border)
    bank_str = f"{clean(bank.get('bank_name'))} | A/c: {clean(bank.get('acc_no'))} | IFSC: {clean(bank.get('ifsc'))}"
    draw_box(ws, f'C5:{last_col_let}5', bank_str, align=left, border=box_border)

    curr_row = 6
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=curr_row, column=i, value=h)
        c.font = Font(bold=True); c.alignment = center; c.border = box_border
        ws.column_dimensions[get_column_letter(i)].width = w
    
    curr = 7
    for item in items:
        for i, key in enumerate(keys, 1):
            val = clean(item.get(key))
            c = ws.cell(row=curr, column=i, value=val); c.border = box_border
            if key in ["quantity", "rate", "gross_amount", "discount_amount", "amount"]: c.alignment = right
            elif key == "particulars": c.alignment = left
            else: c.alignment = center
        curr += 1
    
    fill_to = curr + (12 - len(items))
    while curr < fill_to:
        for i in range(1, num_cols + 1): ws.cell(row=curr, column=i).border = box_border
        curr += 1

    draw_box(ws, f'A{curr}:{get_column_letter(num_cols-1)}{curr}', "Total Amount (Inc. GST)", font=Font(bold=True), align=right, border=box_border)
    total_cell = ws.cell(row=curr, column=num_cols, value=clean(foot.get("total_amount")))
    total_cell.font = Font(bold=True); total_cell.border = box_border; total_cell.alignment = right
    
    curr += 1
    ws.row_dimensions[curr].height = 25
    draw_box(ws, f'A{curr}:{last_col_let}{curr}', f"Amount in Words: {clean(foot.get('amount_in_words'))}", font=Font(italic=True, bold=True), align=left, border=box_border)
    
    curr += 1
    ws.row_dimensions[curr].height = 60
    sig_start = num_cols - 2 if num_cols > 3 else 1
    sig_range = f'{get_column_letter(sig_start)}{curr}:{last_col_let}{curr}'
    draw_box(ws, sig_range, "Authorized Signature", font=Font(bold=True), align=Alignment(horizontal='right', vertical='bottom'), border=box_border)
    for col in range(1, sig_start): ws.cell(row=curr, column=col).border = box_border

    set_outer_border(ws, f'A1:{last_col_let}{curr}')

def write_personal_layout(ws, data):
    # This is the clean layout from Image 1
    items, foot = data.get("items", []), data.get("footer", {})
    ws['A1'] = "EXPENSE SHEET"; ws['A1'].font = Font(size=16, bold=True)
    headers = ["Description", "Quantity", "Rate", "Amount"]
    widths = [40, 10, 15, 15]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=3, column=i, value=h); c.font = Font(bold=True)
        c.border = Border(bottom=Side(style='thin'))
        ws.column_dimensions[get_column_letter(i)].width = w
    curr = 4
    for item in items:
        ws.cell(row=curr, column=1, value=clean(item.get("particulars")))
        ws.cell(row=curr, column=2, value=item.get("quantity")).alignment = Alignment(horizontal='right')
        ws.cell(row=curr, column=3, value=item.get("rate")).alignment = Alignment(horizontal='right')
        ws.cell(row=curr, column=4, value=item.get("amount")).alignment = Alignment(horizontal='right')
        curr += 1
    ws.cell(row=curr+1, column=3, value="TOTAL").font = Font(bold=True); ws.cell(row=curr+1, column=3).alignment = Alignment(horizontal='right')
    ws.cell(row=curr+1, column=4, value=foot.get("total_amount")).font = Font(bold=True); ws.cell(row=curr+1, column=4).alignment = Alignment(horizontal='right')

# ==============================================================================
# 6. ROUTES
# ==============================================================================
@app.after_request
def add_header(response):
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    return response

@app.route("/")
def home(): return render_template("index.html", user=current_user)

@app.route("/login", methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated: return redirect(url_for('home'))
    if request.method == 'POST':
        u = User.query.filter_by(email=request.form.get('email')).first()
        if u and check_password_hash(u.password, request.form.get('password')):
            login_user(u); return redirect(url_for('input_page'))
    return render_template('login.html')

@app.route("/signup", methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        new_u = User(name=request.form.get('name'), email=request.form.get('email'), gender=request.form.get('gender'),
                     password=generate_password_hash(request.form.get('password'), method='scrypt'))
        db.session.add(new_u); db.session.commit()
        return redirect(url_for('login'))
    return render_template('signup.html')

@app.route("/logout")
@login_required
def logout(): logout_user(); return redirect(url_for('home'))

@app.route("/input")
def input_page(): 
    usage = session.get('usage_count', 0)
    if current_user.is_authenticated:
        return render_template("input.html", user=current_user, trials_left=None)
    return render_template("input.html", user=None, trials_left=3-usage)

@app.route("/process", methods=["POST"])
def process():
    if not current_user.is_authenticated and session.get('usage_count', 0) >= 3:
        return jsonify({"error": "3 trials ended, please log in"}), 403
    
    file = request.files.get("image")
    prompt_text = request.form.get("prompt", "")
    
    img_path = os.path.join(UPLOAD_DIR, secure_filename(file.filename) if file else "prompt.png")
    if file: file.save(img_path)
    else: Image.new('RGB', (500, 500), color='white').save(img_path)

    try:
        data = parse_invoice_vision(img_path, prompt_text)
        wb = Workbook(); ws = wb.active
        # The crucial switch based on the AI's determination
        if data.get("layout") == "personal": write_personal_layout(ws, data)
        else: write_business_layout(ws, data)
        
        save_path = os.path.join(UPLOAD_DIR, "Data.xlsx"); wb.save(save_path)
        global LAST_GENERATED_FILE; LAST_GENERATED_FILE = save_path
        
        if not current_user.is_authenticated: session['usage_count'] = session.get('usage_count', 0) + 1
        return jsonify({"status": "ok", "filename": "Data.xlsx", "trials_left": 3-session.get('usage_count', 0)})
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/download")
def download():
    path = os.path.join(UPLOAD_DIR, secure_filename(request.args.get('filename', 'Data.xlsx')))
    return send_file(path, as_attachment=True)

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host='0.0.0.0', port=port)
